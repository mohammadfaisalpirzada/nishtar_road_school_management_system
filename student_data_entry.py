#!/usr/bin/env python3
"""
Student Data Entry Program
A CLI-based program to add student data to Excel file
"""

import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime
import re
import msvcrt  # For Windows keyboard input

class StudentDataEntry:
    def __init__(self, excel_file="students_data_2025.xlsx"):
        self.excel_file = excel_file
        self.headers = [
            "Class_S.No", 
            "GR#",
            "Student Name",
            "Father's Name",
            "Gender",
            "Religion",
            "Contact Number",
            "CNIC / B-Form",
            "Date of Birth",
            "Father/Mother's CNIC",
            "Guardian Name",
            "Guardian CNIC",
            "Guardian Relation",
            "Student Class",
            "Class Section",
            "SEMIS Code",
            "Date of Admission"
        ]
        self.setup_excel_file()
    
    def setup_excel_file(self):
        """Setup Excel file with headers if it doesn't exist or is empty"""
        try:
            if os.path.exists(self.excel_file):
                self.workbook = openpyxl.load_workbook(self.excel_file)
                # Use specific worksheet '408070227' as default
                if '408070227' in self.workbook.sheetnames:
                    self.worksheet = self.workbook['408070227']
                else:
                    # Create the worksheet if it doesn't exist
                    self.worksheet = self.workbook.create_sheet('408070227')
                    self.add_headers_to_sheet(self.worksheet)
                
                # Check if headers exist
                if self.worksheet.max_row == 1 and self.worksheet['A1'].value is None:
                    self.add_headers_to_sheet(self.worksheet)
            else:
                self.workbook = Workbook()
                self.worksheet = self.workbook.active
                self.worksheet.title = '408070227'
                self.add_headers_to_sheet(self.worksheet)
                
        except Exception as e:
            print(f"Error setting up Excel file: {e}")
            # Create new workbook if there's an issue
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = '408070227'
            self.add_headers_to_sheet(self.worksheet)
    
    def add_headers(self):
        """Add headers to the current worksheet"""
        self.add_headers_to_sheet(self.worksheet)
    
    def add_headers_to_sheet(self, worksheet):
        """Add headers to a specific worksheet"""
        for col, header in enumerate(self.headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        self.save_file()
    
    def get_or_create_class_sheet(self, student_class):
        """Get or create a worksheet for a specific class"""
        sheet_name = f"Class_{student_class}"
        
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        else:
            # Create new sheet for this class
            new_sheet = self.workbook.create_sheet(sheet_name)
            self.add_headers_to_sheet(new_sheet)
            return new_sheet
    
    def check_duplicate_gr(self, gr_number):
        """Check if GR# already exists in any worksheet"""
        try:
            # Check all sheets for duplicate GR# (column 3)
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                for row in range(2, sheet.max_row + 1):
                    existing_gr = sheet.cell(row=row, column=3).value
                    if existing_gr and str(existing_gr).strip() == str(gr_number).strip():
                        return True
            return False
        except Exception:
            return False
    
    def format_cnic(self, cnic_number):
        """Format CNIC number by adding dashes in correct positions"""
        # Don't format N/A values
        if cnic_number == "N/A":
            return cnic_number
            
        # Remove any existing dashes and spaces
        clean_cnic = re.sub(r'[^0-9]', '', cnic_number)
        
        # Check if it's a valid length (13 digits)
        if len(clean_cnic) == 13:
            # Format as XXXXX-XXXXXXX-X
            return f"{clean_cnic[:5]}-{clean_cnic[5:12]}-{clean_cnic[12]}"
        else:
            # Return as-is if not 13 digits
            return cnic_number
    
    def validate_input(self, field_name, value):
        """Validate input based on field type"""
        if not value.strip():
            return False, "This field cannot be empty"
        
        # Allow N/A for optional fields
        if value.strip() == "N/A":
            return True, ""
        
        # Specific validations
        if field_name == "GR#":
            # Check if it's numeric only
            if not value.isdigit():
                return False, "GR# must contain numbers only (no letters or symbols)"
            
            # Check for duplicates
            if self.check_duplicate_gr(value):
                return False, "This GR# already exists. Please enter a unique GR#"
        
        elif field_name == "Contact Number":
            if value != "N/A" and not re.match(r'^[0-9+\-\s()]+$', value):
                return False, "Invalid phone number format"
        
        elif field_name in ["CNIC / B-Form", "Father/Mother's CNIC", "Guardian CNIC"]:
            if value != "N/A":
                # Remove dashes and check if it's all digits
                clean_value = re.sub(r'[^0-9]', '', value)
                if not clean_value.isdigit():
                    return False, "CNIC must contain only numbers"
                if len(clean_value) != 13:
                    return False, "CNIC must be exactly 13 digits"
        
        elif field_name == "Gender":
            if value.lower() not in ['male', 'female', 'm', 'f']:
                return False, "Gender must be Male/Female or M/F"
        
        elif field_name in ["Date of Birth", "Date of Admission"]:
            if value != "N/A":
                try:
                    # Try to parse date in various formats
                    for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d.%m.%Y']:
                        try:
                            datetime.strptime(value, fmt)
                            return True, ""
                        except ValueError:
                            continue
                    return False, "Invalid date format. Use DD/MM/YYYY or DD-MM-YYYY"
                except:
                    return False, "Invalid date format"
        
        return True, ""
    
    def get_next_class_serial_number(self, student_class):
        """Get the next class-specific serial number for a given class"""
        sheet_name = f"Class_{student_class}"
        
        # If sheet doesn't exist, this will be the first student
        if sheet_name not in self.workbook.sheetnames:
            return 1
        
        sheet = self.workbook[sheet_name]
        max_class_sno = 0
        
        # Check the specific class sheet for the highest Class_S.No
        for row in range(2, sheet.max_row + 1):
            # Check if the row has any data
            if (sheet.cell(row=row, column=1).value is not None or
                sheet.cell(row=row, column=2).value is not None or
                sheet.cell(row=row, column=3).value is not None):
                
                # Get the Class_S.No from this row (column 1)
                class_sno = sheet.cell(row=row, column=1).value
                if class_sno is not None:
                    try:
                        # Extract number from different formats
                        if isinstance(class_sno, str):
                            if '_' in class_sno:  # Format like "ECE_01"
                                sno_part = class_sno.split('_')[-1]
                                sno_int = int(sno_part)
                            elif '-' in class_sno:  # Format like "ECE-01" (legacy)
                                sno_part = class_sno.split('-')[-1]
                                sno_int = int(sno_part)
                            elif class_sno.isdigit() and len(class_sno) >= 4:  # Format like "1001", "2001", etc.
                                # Extract last three digits for numbered classes (class prefix + serial number)
                                sno_int = int(class_sno[-3:])
                            elif class_sno.isdigit() and len(class_sno) == 3:  # Legacy format like "101", "201", etc.
                                # Extract last two digits for legacy numbered classes
                                sno_int = int(class_sno[-2:])
                            else:
                                sno_int = int(class_sno)
                        else:
                            sno_int = int(class_sno)
                        
                        if sno_int > max_class_sno:
                            max_class_sno = sno_int
                    except (ValueError, TypeError):
                        pass
        
        return max_class_sno + 1
    

    
    def get_total_students(self):
        """Get total number of students across all class sheets"""
        try:
            total = 0
            class_names = ['ECE', 'I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X']
            
            for class_name in class_names:
                sheet_name = f"Class_{class_name}"
                if sheet_name in self.workbook.sheetnames:
                    sheet = self.workbook[sheet_name]
                    # Count rows with data (excluding header)
                    for row in range(2, sheet.max_row + 1):
                        if sheet.cell(row=row, column=1).value is not None:
                            total += 1
            
            return total
            
        except Exception as e:
            print(f"Error getting total students: {e}")
            return 0
    
    def get_class_student_count(self, class_name):
        """Get number of students in a specific class"""
        try:
            # Use the correct sheet naming convention
            sheet_name = f"Class_{class_name}"
            if sheet_name not in self.workbook.sheetnames:
                return 0
            
            sheet = self.workbook[sheet_name]
            count = 0
            
            # Count rows with data (excluding header)
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=1).value is not None:
                    count += 1
            
            return count
        except Exception as e:
            print(f"Error getting class student count: {e}")
            return 0
    
    def get_class_section_count(self, class_name, section):
        """Get number of students in a specific class and section"""
        try:
            # Use the correct sheet naming convention
            sheet_name = f"Class_{class_name}" if class_name != "ECE" else "ECE"
            if sheet_name not in self.workbook.sheetnames:
                return 0
            
            sheet = self.workbook[sheet_name]
            count = 0
            
            # Find the Class Section column (usually column 15)
            section_col = None
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == "Class Section":
                    section_col = col
                    break
            
            if section_col is None:
                return 0
            
            # Count rows with matching section
            for row in range(2, sheet.max_row + 1):
                if (sheet.cell(row=row, column=1).value is not None and
                    sheet.cell(row=row, column=section_col).value == section):
                    count += 1
            
            return count
        except Exception as e:
            print(f"Error getting class section count: {e}")
            return 0
    
    def get_class_gender_count(self, class_name, gender):
        """Get number of students in a specific class by gender"""
        try:
            # Use the correct sheet naming convention
            sheet_name = f"Class_{class_name}"
            if sheet_name not in self.workbook.sheetnames:
                return 0
            
            sheet = self.workbook[sheet_name]
            count = 0
            
            # Find the Gender column (usually column 5)
            gender_col = None
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == "Gender":
                    gender_col = col
                    break
            
            if gender_col is None:
                return 0
            
            # Count rows with matching gender
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=1).value is not None:
                    student_gender = sheet.cell(row=row, column=gender_col).value
                    if student_gender and student_gender.strip().lower() == gender.lower():
                        count += 1
            
            return count
        except Exception as e:
            print(f"Error getting class gender count: {e}")
            return 0
    
    def collect_student_data(self):
        """Collect student data from user input"""
        print("\n" + "="*60)
        print("           STUDENT DATA ENTRY SYSTEM")
        print("="*60)
        
        student_data = {}
        
        # Collect other data
        for header in self.headers:  # Process all headers
            # Skip auto-generated fields initially
            if header in ["Class_S.No", "SEMIS Code", "Class Section"]:
                continue
            
            # Skip Guardian CNIC and Relation if Guardian Name is already set to "-"
            if header in ["Guardian CNIC", "Guardian Relation"] and student_data.get("Guardian Name") == "-":
                continue
            
            # Skip guardian fields if they're already set from father's data
            if header in ["Guardian Name", "Guardian CNIC", "Guardian Relation"] and student_data.get("Guardian Name") and student_data.get("Guardian Name") != "":
                continue
            
            # Special handling for Guardian Name - ask for guardian type first
            if header == "Guardian Name":
                while True:
                    print("\nGuardian Type: (F=Father, O=Others, N=Nil)", end=" ")
                    guardian_type = input().strip().upper()
                    
                    if guardian_type == "F":
                        # Copy father's data to guardian fields
                        father_name = student_data.get("Father's Name", "")
                        father_cnic = student_data.get("Father/Mother's CNIC", "")
                        contact_number = student_data.get("Contact Number", "")
                        
                        student_data["Guardian Name"] = father_name
                        student_data["Guardian CNIC"] = father_cnic
                        student_data["Guardian Relation"] = "Father"
                        
                        print(f"   Guardian Name: {father_name} (Copied from Father's Name)")
                        print(f"   Guardian CNIC: {father_cnic} (Copied from Father's CNIC)")
                        print(f"   Guardian Relation: Father (Auto-set)")
                        break
                    
                    elif guardian_type == "N":
                        # Set all guardian fields to "-"
                        student_data["Guardian Name"] = "-"
                        student_data["Guardian CNIC"] = "-"
                        student_data["Guardian Relation"] = "-"
                        
                        print(f"   Guardian Name: - (Auto-set)")
                        print(f"   Guardian CNIC: - (Auto-set)")
                        print(f"   Guardian Relation: - (Auto-set)")
                        break
                    
                    elif guardian_type == "O":
                        # Continue with normal guardian data collection
                        break
                    
                    else:
                        print("   ❌ Error: Please enter F (Father), O (Others), or N (Nil)")
                
                # If guardian type is "O", continue with normal collection
                if guardian_type != "O":
                    continue
                
            while True:
                print(f"\n{header}:", end=" ")
                
                # Provide hints for certain fields
                if header == "Gender":
                    print("(Male/Female or M/F)", end=" ")
                elif header in ["Date of Birth", "Date of Admission"]:
                    print("(DD/MM/YYYY or DD-MM-YYYY)", end=" ")
                elif header == "Contact Number":
                    print("(e.g., +92-300-1234567)", end=" ")
                elif header in ["CNIC / B-Form", "Father/Mother's CNIC", "Guardian CNIC"]:
                    print("(e.g., 12345-6789012-3)", end=" ")
                
                value = input().strip()
                
                # Allow empty for optional fields with N/A default
                optional_fields = ["Guardian Name", "Guardian CNIC", "Guardian Relation", "Contact Number", "CNIC / B-Form", "Date of Birth", "Father/Mother's CNIC", "Date of Admission"]
                if not value and header in optional_fields:
                    student_data[header] = "N/A"
                    break
                
                is_valid, error_msg = self.validate_input(header, value)
                if is_valid:
                    # Format gender
                    if header == "Gender":
                        value = "Male" if value.lower() in ['male', 'm'] else "Female"
                    
                    # Format CNIC numbers automatically
                    elif header in ["CNIC / B-Form", "Father/Mother's CNIC", "Guardian CNIC"]:
                        value = self.format_cnic(value)
                    
                    student_data[header] = value
                    
                    # Generate Class_S.No after Student Class is collected
                    if header == "Student Class":
                        class_sno = self.get_next_class_serial_number(student_data["Student Class"])
                        student_class = student_data["Student Class"]
                        
                        # Format based on class type
                        if student_class == "ECE":
                            student_data["Class_S.No"] = f"ECE_{class_sno:02d}"
                        else:
                            # For numbered classes (I, II, III, etc.), use format like 101, 201, 301
                            class_mapping = {
                                "I": 1, "II": 2, "III": 3, "IV": 4, "V": 5,
                                "VI": 6, "VII": 7, "VIII": 8, "IX": 9, "X": 10
                            }
                            if student_class in class_mapping:
                                class_prefix = class_mapping[student_class]
                                student_data["Class_S.No"] = f"{class_prefix}{class_sno:03d}"
                            else:
                                # Fallback for any other class format
                                student_data["Class_S.No"] = f"{student_class}_{class_sno:02d}"
                        
                        print(f"\nClass S.No: {student_data['Class_S.No']} (Auto-generated)")
                    
                    break
                else:
                    print(f"   ❌ Error: {error_msg}. Please try again.")
        
        # Auto-set SEMIS Code and Class Section based on gender
        student_data["SEMIS Code"] = "408070227"
        gender = student_data.get("Gender", "").lower()
        if gender in ["male", "m"]:
            student_data["Class Section"] = "Boys"
        else:  # female or any other value defaults to Girls
            student_data["Class Section"] = "Girls"
        print(f"\n   SEMIS Code: 408070227 (Auto-set)")
        print(f"   Class Section: {student_data['Class Section']} (Auto-set based on gender)")
        
        return student_data
    
    def add_student_record(self, student_data):
        """Add student record to Excel file in class-specific sheet"""
        try:
            # Get the student's class
            student_class = student_data.get('Student Class', '')
            
            # Get or create the appropriate class sheet
            if student_class:
                target_sheet = self.get_or_create_class_sheet(student_class)
            else:
                # Use default sheet if no class specified
                target_sheet = self.worksheet
            
            # Find the last row with actual data in the target sheet
            last_data_row = 1  # Start with header row
            for row in range(2, target_sheet.max_row + 1):
                # Check if the row has any data in the first few columns
                if (target_sheet.cell(row=row, column=1).value is not None or
                    target_sheet.cell(row=row, column=3).value is not None or
                    target_sheet.cell(row=row, column=4).value is not None):
                    last_data_row = row
            
            # Add data to the next row after last data
            next_row = last_data_row + 1
            
            for col, header in enumerate(self.headers, 1):
                target_sheet.cell(row=next_row, column=col, value=student_data[header])
            
            self.save_file()
            return True
        except Exception as e:
            print(f"Error adding record: {e}")
            return False
    
    def save_file(self):
        """Save the Excel file"""
        try:
            self.workbook.save(self.excel_file)
        except Exception as e:
            print(f"Error saving file: {e}")
    
    def display_summary(self, student_data):
        """Display summary of entered data"""
        print("\n" + "="*60)
        print("           DATA ENTRY SUMMARY")
        print("="*60)
        
        for header in self.headers:
            value = student_data.get(header, "")
            print(f"{header:<25}: {value}")
        
        print("="*60)
    
    def confirm_entry(self, student_data):
        """Confirm data entry with user"""
        self.display_summary(student_data)
        
        while True:
            choice = input("\nDo you want to save this record? (y/n): ").lower().strip()
            if choice in ['y', 'yes']:
                return True
            elif choice in ['n', 'no']:
                return False
            else:
                print("Please enter 'y' for yes or 'n' for no.")
    
    def run(self):
        """Main program loop"""
        print("Welcome to Student Data Entry System!")
        print(f"Working with file: {self.excel_file}")
        print("\n📝 Instructions:")
        print("   • Press ENTER after completing a student entry to add another")
        print("   • Press ESC to exit the program")
        
        while True:
            try:
                # Collect student data
                student_data = self.collect_student_data()
                
                # Confirm entry
                if self.confirm_entry(student_data):
                    if self.add_student_record(student_data):
                        print("\n✅ Student record saved successfully!")
                    else:
                        print("\n❌ Failed to save student record.")
                        continue
                else:
                    print("\n❌ Record not saved.")
                    continue
                
                # Ask for continuation with Enter/Escape
                print("\n" + "="*60)
                print("Press ENTER to add another student or ESC to exit...")
                
                while True:
                    key = msvcrt.getch()
                    if key == b'\r':  # Enter key
                        break
                    elif key == b'\x1b':  # Escape key
                        print("\nThank you for using Student Data Entry System!")
                        return
                        
            except KeyboardInterrupt:
                print("\n\nProgram interrupted by user. Goodbye!")
                break
            except Exception as e:
                print(f"\nAn error occurred: {e}")
                print("Please try again.")

def main():
    """Main function"""
    app = StudentDataEntry()
    app.run()

if __name__ == "__main__":
    main()