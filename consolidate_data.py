#!/usr/bin/env python3
"""
Data Consolidation Program
Collects data from all sheets in students_data_2025.xlsx and creates a consolidated file 408070227.xlsx
"""

import openpyxl
from openpyxl import Workbook
import os
import re
from datetime import datetime

def format_cnic(cnic_value):
    """
    Format CNIC/B-Form number to proper format (XXXXX-XXXXXXX-X)
    """
    if not cnic_value or str(cnic_value).strip() in ['', '-', 'N/A', 'None']:
        return 'N/A'
    
    # Remove all non-digit characters
    digits_only = re.sub(r'\D', '', str(cnic_value))
    
    # Check if we have 13 digits for CNIC
    if len(digits_only) == 13:
        return f"{digits_only[:5]}-{digits_only[5:12]}-{digits_only[12]}"
    # Check if we have valid B-Form format (varies)
    elif len(digits_only) >= 10:
        return digits_only
    else:
        return 'N/A'

def format_mobile_number(mobile_value):
    """
    Format mobile number to proper format
    """
    if not mobile_value or str(mobile_value).strip() in ['', '-', 'N/A', 'None']:
        return 'N/A'
    
    # Remove all non-digit characters
    digits_only = re.sub(r'\D', '', str(mobile_value))
    
    # Check for valid Pakistani mobile number (11 digits starting with 03)
    if len(digits_only) == 11 and digits_only.startswith('03'):
        return f"{digits_only[:4]}-{digits_only[4:]}"
    # Check for 10 digit number (add 0 prefix)
    elif len(digits_only) == 10 and digits_only.startswith('3'):
        return f"0{digits_only[:3]}-{digits_only[3:]}"
    # Return as is if valid length
    elif len(digits_only) >= 10:
        return digits_only
    else:
        return 'N/A'

def format_date(date_value):
    """
    Format date to dd-mmm-yyyy format (e.g., 15-Jan-2025)
    """
    if not date_value or str(date_value).strip() in ['', '-', 'N/A', 'None']:
        return 'N/A'
    
    try:
        # If it's already a datetime object
        if isinstance(date_value, datetime):
            return date_value.strftime('%d-%b-%Y')
        
        # Handle Excel serial date numbers (like 42865)
        if isinstance(date_value, (int, float)) or str(date_value).strip().isdigit():
            try:
                # Excel serial date starts from 1900-01-01 (serial number 1)
                # But Excel incorrectly treats 1900 as a leap year, so we adjust
                serial_num = int(float(str(date_value).strip()))
                if 1 <= serial_num <= 2958465:  # Valid Excel date range
                    # Convert Excel serial to datetime
                    # Excel epoch is 1900-01-01, but we need to account for the leap year bug
                    if serial_num > 59:  # After Feb 28, 1900
                        serial_num -= 1  # Adjust for Excel's leap year bug
                    
                    from datetime import date, timedelta
                    excel_epoch = date(1900, 1, 1)
                    converted_date = excel_epoch + timedelta(days=serial_num - 1)
                    return converted_date.strftime('%d-%b-%Y')
            except (ValueError, OverflowError):
                pass
        
        # Convert string to datetime and format
        date_str = str(date_value).strip()
        
        # Try different date formats
        date_formats = [
            '%Y-%m-%d',      # 2025-01-15
            '%d/%m/%Y',      # 15/01/2025
            '%m/%d/%Y',      # 01/15/2025
            '%d-%m-%Y',      # 15-01-2025
            '%d.%m.%Y',      # 15.01.2025
            '%Y/%m/%d',      # 2025/01/15
            '%d %m %Y',      # 15 01 2025
            '%d-%b-%Y',      # 15-Jan-2025 (already formatted)
            '%d %b %Y',      # 15 Jan 2025
            '%b %d, %Y',     # Jan 15, 2025
            '%B %d, %Y',     # January 15, 2025
        ]
        
        for fmt in date_formats:
            try:
                parsed_date = datetime.strptime(date_str, fmt)
                return parsed_date.strftime('%d-%b-%Y')
            except ValueError:
                continue
        
        # If no format matches, return as is
        return date_str
        
    except Exception:
        return 'N/A'

def clean_data_value(value):
    """
    Clean data value - replace empty, None, or '-' with 'N/A'
    """
    if value is None or str(value).strip() in ['', '-', 'None']:
        return 'N/A'
    return str(value).strip()

def consolidate_student_data():
    """
    Consolidate all student data from students_data_2025.xlsx into 408070227.xlsx
    """
    
    # Define the source and target files
    source_file = 'students_data_2025.xlsx'
    target_file = '408070227.xlsx'
    
    # Check if source file exists
    if not os.path.exists(source_file):
        print(f"Error: {source_file} not found!")
        return
    
    try:
        # Load the source workbook
        print(f"Loading {source_file}...")
        source_wb = openpyxl.load_workbook(source_file)
        
        # Create new workbook for consolidated data
        target_wb = Workbook()
        target_ws = target_wb.active
        target_ws.title = "Consolidated_Data"
        
        # Define class order: ECE, I, II, III, IV, V, VI, VII, VIII, IX, X
        class_order = ['ECE', 'I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X']
        
        # Set up headers in target sheet
        headers = [
            'S.No',  # Combined S.No (new first column)
            'Class_S.No',
            'GR#',
            'Student Name',
            'Father\'s Name',
            'Gender',
            'Religion',
            'Contact Number',
            'CNIC / B-Form',
            'Date of Birth',
            'Father/Mother\'s CNIC',
            'Guardian Name',
            'Guardian CNIC',
            'Guardian Relation',
            'Student Class',
            'Class Section',
            'SEMIS Code',
            'Date of Admission'
        ]
        
        # Write headers to target sheet
        for col, header in enumerate(headers, 1):
            target_ws.cell(row=1, column=col, value=header)
        
        # Initialize combined S.No counter
        combined_sno = 1
        target_row = 2
        
        print("Consolidating data from all classes...")
        
        # Process each class in order
        for class_name in class_order:
            sheet_name = f"Class_{class_name}"
            
            if sheet_name in source_wb.sheetnames:
                print(f"Processing {sheet_name}...")
                source_ws = source_wb[sheet_name]
                
                # Get header row from source to map columns
                source_headers = {}
                for col in range(1, source_ws.max_column + 1):
                    header_value = source_ws.cell(row=1, column=col).value
                    if header_value:
                        source_headers[header_value] = col
                
                # Process each student row in the source sheet
                for row in range(2, source_ws.max_row + 1):
                    # Check if row has data (S.No exists)
                    if source_ws.cell(row=row, column=1).value:
                        # Write combined S.No in first column
                        target_ws.cell(row=target_row, column=1, value=combined_sno)
                        
                        # Copy data from source to target (starting from column 2)
                        col_mapping = {
                            'Class_S.No': 2,
                            'GR#': 3,
                            'Student Name': 4,
                            'Father\'s Name': 5,
                            'Gender': 6,
                            'Religion': 7,
                            'Contact Number': 8,
                            'CNIC / B-Form': 9,
                            'Date of Birth': 10,
                            'Father/Mother\'s CNIC': 11,
                            'Guardian Name': 12,
                            'Guardian CNIC': 13,
                            'Guardian Relation': 14,
                            'Student Class': 15,
                            'Class Section': 16,
                            'SEMIS Code': 17,
                            'Date of Admission': 18
                        }
                        
                        # Copy each field with proper formatting
                        for field, target_col in col_mapping.items():
                            if field in source_headers:
                                source_col = source_headers[field]
                                value = source_ws.cell(row=row, column=source_col).value
                                
                                # Apply specific formatting based on field type
                                if field in ['CNIC / B-Form', 'Father/Mother\'s CNIC', 'Guardian CNIC']:
                                    formatted_value = format_cnic(value)
                                elif field == 'Contact Number':
                                    formatted_value = format_mobile_number(value)
                                elif field in ['Date of Birth', 'Date of Admission']:
                                    formatted_value = format_date(value)
                                else:
                                    formatted_value = clean_data_value(value)
                                
                                target_ws.cell(row=target_row, column=target_col, value=formatted_value)
                            else:
                                # If field not found in source, set as N/A
                                target_ws.cell(row=target_row, column=target_col, value='N/A')
                        
                        combined_sno += 1
                        target_row += 1
            else:
                print(f"Warning: {sheet_name} not found in source file")
        
        # Save the consolidated file (overwrite if exists)
        print(f"Saving consolidated data to {target_file}...")
        target_wb.save(target_file)
        
        print(f"\n✅ Consolidation completed successfully!")
        print(f"📊 Total students consolidated: {combined_sno - 1}")
        print(f"📁 Output file: {target_file}")
        
    except Exception as e:
        print(f"❌ Error during consolidation: {str(e)}")
    
    finally:
        # Close workbooks
        if 'source_wb' in locals():
            source_wb.close()
        if 'target_wb' in locals():
            target_wb.close()

if __name__ == "__main__":
    print("=" * 50)
    print("📚 STUDENT DATA CONSOLIDATION PROGRAM")
    print("=" * 50)
    print("This program consolidates all student data from students_data_2025.xlsx")
    print("into a single file named 408070227.xlsx with combined S.No")
    print("\nClass order: ECE → I → II → III → IV → V → VI → VII → VIII → IX → X")
    print("-" * 50)
    
    consolidate_student_data()
    
    print("-" * 50)
    print("Program execution completed.")
    input("Press Enter to exit...")